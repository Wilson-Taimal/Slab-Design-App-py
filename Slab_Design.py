from tkinter import *
vent = Tk()
vent.geometry("995x845")
vent.title(" SLAB DESIGN ")
vent.iconbitmap('D:\\BIBLIOTECA PERSONAL\\Programación\\Python\\logo-wat.ico')

from Ecuaciones import *

# Función calcular
def Calcular():
    # Materiales
    fc = float (c1ent01.get())
    fy = float (c1ent02.get())
    b = float (c1ent03.get())
    h = float (c1ent04.get())
    pmin =  float (c1ent05.get())
    fiv = float (c1ent06.get())

    # Datos iniciales Parrilla sup
    rs = float (c1ent07.get())
    ds = h-rs
    Mnx = float (c2ent04.get())
    Mny = float (c2ent54.get())
    Nbsx = float (c2ent08.get())
    Nbsy = float (c2ent58.get())

    # Datos iniciales Parrilla inf
    ri = float (c1ent08.get())
    di = h-ri
    Mpx = float (c3ent04.get())
    Mpy = float (c3ent54.get())
    Nbix = float (c3ent08.get())
    Nbiy = float (c3ent58.get())

    # Datos iniciales diseño cortante simple
    r = max(rs, ri)
    d = h-r
    Vux = float (c4ent04.get())
    Vuy = float (c4ent54.get())

    # Acero minimo
    Asmin = f_Asmin(pmin, b, h); print("Asmin = %.2f" %Asmin)
    c1ent09.delete(0, 'end');    c1ent09.insert (0, "{:.2f}".format(Asmin))

    #Parrilla superior dir x - Diseño
    print(''); print('Parilla superior dir x')
    pcalsx = f_pcal(fc, fy, Mnx, b, ds);        print(" pcal  = %.4f" %pcalsx)
    c2ent05.delete(0, 'end');                   c2ent05.insert (0, "{:.4f}".format(pcalsx))
       
    Ascalsx = f_Ascal(pcalsx, b, ds);           print(" Ascal = %.2f" %Ascalsx)
    c2ent06.delete(0, 'end');                   c2ent06.insert (0, "{:.2f}".format(Ascalsx))
    
    Asreqsx = f_Asreq(Ascalsx, Asmin);          print(" Asreq = %.2f" %Asreqsx)
    c2ent07.delete(0, 'end');                   c2ent07.insert (0, "{:.2f}".format(Asreqsx))
    
    Asbsx = f_Asb(Nbsx);                        print(" Asb   = %.2f" %Asbsx)
    Sepsx = f_sep(Asreqsx, Asbsx, b);           print(" Sep   = %.2f" %Sepsx)

    dbpsx = f_dbp(Nbsx)
    Dissx = f_dist(Asreqsx, Asbsx, b, dbpsx);    print(' Ref   = ', Dissx)
    c2ent09.delete(0, 'end');                   c2ent09.insert (0, Dissx)
    
    Ascolsx = f_Ascol(Asreqsx, Asbsx);          print(" Ascol = %.2f" %Ascolsx)
    c2ent13.delete(0, 'end');                   c2ent13.insert (0, "{:.2f}".format(Ascolsx))
    
    Mnsx = f_Mn(fc, fy, Ascolsx, b, ds);        print(" Mn    = %.1f" %Mnsx)
    c2ent14.delete(0, 'end');                   c2ent14.insert (0, "{:.0f}".format(Mnsx))
    
    ChMnsx = f_ChMn(Mnx,Mnsx);                  print(' Chequeo Mn =', ChMnsx)
    c2ent15.delete(0, 'end');                   c2ent15.insert (0, ChMnsx)

    #Parrilla superior dir y - Diseño
    print(''); print('Parilla superior dir y')
    pcalsy = f_pcal(fc, fy, Mny, b, ds);        print(" pcal  = %.4f" %pcalsy)
    c2ent55.delete(0, 'end');                   c2ent55.insert (0, "{:.4f}".format(pcalsy))

    Ascalsy = f_Ascal(pcalsy, b, ds);           print(" Ascal = %.2f" %Ascalsy)
    c2ent56.delete(0, 'end');                   c2ent56.insert (0, "{:.2f}".format(Ascalsy))

    Asreqsy = f_Asreq(Ascalsy, Asmin);          print(" Asreq = %.2f" %Asreqsy)
    c2ent57.delete(0, 'end');                   c2ent57.insert (0, "{:.2f}".format(Asreqsy))

    Asbsy = f_Asb(Nbsy);                        print(" Asb   = %.2f" %Asbsy)
    Sepsy = f_sep(Asreqsy, Asbsy, b);           print(" Sep   = %.2f" %Sepsy)

    dbpsy = f_dbp(Nbsy)                        
    Dissy = f_dist(Asreqsy, Asbsy, b, dbpsy);    print(' Ref   = ', Dissy)
    c2ent59.delete(0, 'end');                   c2ent59.insert (0, Dissy)

    Ascolsy = f_Ascol(Asreqsy, Asbsy);          print(" Ascol = %.2f" %Ascolsy)
    c2ent63.delete(0, 'end');                   c2ent63.insert (0, "{:.2f}".format(Ascolsy))

    Mnsy = f_Mn(fc, fy, Ascolsy, b, ds);        print(" Mn    = %.1f" %Mnsy)
    c2ent64.delete(0, 'end');                   c2ent64.insert (0, "{:.0f}".format(Mnsy))

    ChMnsy = f_ChMn(Mny,Mnsy);                  print(' Chequeo Mn = ', ChMnsy)
    c2ent65.delete(0, 'end');                   c2ent65.insert (0, ChMnsy)

    #Parrilla inferior dir x - Diseño
    print(''); print('Parilla inferior dir x')
    pcalix = f_pcal(fc, fy, Mpx, b, di);        print(" pcal  = %.4f" %pcalix)
    c3ent05.delete(0, 'end');                   c3ent05.insert (0, "{:.4f}".format(pcalix))

    Ascalix = f_Ascal(pcalix, b, di);           print(" Ascal = %.2f" %Ascalix)
    c3ent06.delete(0, 'end');                   c3ent06.insert (0, "{:.2f}".format(Ascalix))

    Asreqix = f_Asreq(Ascalix, Asmin);          print(" Asreq = %.2f" %Asreqix)
    c3ent07.delete(0, 'end');                   c3ent07.insert (0, "{:.2f}".format(Asreqix))

    Asbix = f_Asb(Nbix);                        print(" Asb   = %.2f" %Asbix)
    Sepix = f_sep(Asreqix, Asbix, b);           print(" Sep   = %.2f" %Sepix)

    dbpix = f_dbp(Nbix)                        
    Disix = f_dist(Asreqix, Asbix, b, dbpix);    print(' Ref   = ', Disix)
    c3ent09.delete(0, 'end');                   c3ent09.insert (0, Disix)

    Ascolix = f_Ascol(Asreqix, Asbix);          print(" Ascol = %.2f" %Ascolix)
    c3ent13.delete(0, 'end');                   c3ent13.insert (0, "{:.2f}".format(Ascolix))

    Mnix = f_Mn(fc, fy, Ascolix, b, di);        print(" Mn    = %.1f" %Mnix)
    c3ent14.delete(0, 'end');                   c3ent14.insert (0, "{:.0f}".format(Mnix))

    ChMnix = f_ChMn(Mpx,Mnix);                  print(' Chequeo Mn = ', ChMnix)
    c3ent15.delete(0, 'end');                   c3ent15.insert (0, ChMnix)

    #Parrilla inferior dir y - Diseño
    print(''); print('Parilla inferior dir y')
    pcaliy = f_pcal(fc, fy, Mpy, b, di);        print(" pcal  = %.4f" %pcaliy)
    c3ent55.delete(0, 'end');                   c3ent55.insert (0, "{:.4f}".format(pcaliy))

    Ascaliy = f_Ascal(pcaliy, b, di);           print(" Ascal = %.2f" %Ascaliy)
    c3ent56.delete(0, 'end');                   c3ent56.insert (0, "{:.2f}".format(Ascaliy))

    Asreqiy = f_Asreq(Ascaliy, Asmin);          print(" Asreq = %.2f" %Asreqiy)
    c3ent57.delete(0, 'end');                   c3ent57.insert (0, "{:.2f}".format(Asreqiy))

    Asbiy = f_Asb(Nbiy);                        print(" Asb   = %.2f" %Asbiy)
    Sepiy = f_sep(Asreqiy, Asbiy, b);           print(" Sep   = %.2f" %Sepiy)

    dbpiy = f_dbp(Nbiy);                        
    Disiy = f_dist(Asreqiy, Asbiy, b, dbpiy);    print(' Ref   = ', Disiy)
    c3ent59.delete(0, 'end');                   c3ent59.insert (0, Disiy)

    Ascoliy = f_Ascol(Asreqiy, Asbiy);          print(" Ascol = %.2f" %Ascoliy)
    c3ent63.delete(0, 'end');                   c3ent63.insert (0, "{:.2f}".format(Ascoliy))

    Mniy = f_Mn(fc, fy, Ascoliy, b, di);        print(" Mn    = %.1f" %Mniy)
    c3ent64.delete(0, 'end');                   c3ent64.insert (0, "{:.0f}".format(Mniy))

    ChMniy = f_ChMn(Mpy,Mniy);                  print(' Chequeo Mn = ', ChMniy)
    c3ent65.delete(0, 'end');                   c3ent65.insert (0, ChMniy)

    #Diseño a cortante
    print(''); print('Diseño a cortante')
    Vcx = f_Vc(fiv, fc, b, d);                  print(" Vcx = %.1f" %Vcx)
    c4ent05.delete(0, 'end');                   c4ent05.insert (0, "{:.2f}".format(Vcx))

    ChVcx = f_ChVc(Vux,Vcx);                    print(' Chequeo cortante = ', ChVcx)
    c4ent06.delete(0, 'end');                   c4ent06.insert (0, ChVcx)

    Vcy = f_Vc(fiv, fc, b, d);                  print(" Vcy = %.1f" %Vcy)
    c4ent55.delete(0, 'end');                   c4ent55.insert (0, "{:.2f}".format(Vcy))

    ChVcy = f_ChVc(Vuy,Vcy);                    print(' Chequeo cortante = ', ChVcy)
    c4ent56.delete(0, 'end');                   c4ent56.insert (0, ChVcy)

# Función guardar
def Guardar():
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font
    book = Workbook()
    book = load_workbook('D:\IEB_LOSAS\PlantillaSlabDesign.xlsx')
    sheet = book.active

    # Materiales
    fc = float (c1ent01.get())
    fy = float (c1ent02.get())
    b = float (c1ent03.get());    sheet['C8'] = b;  sheet['C36'] = b*10;    sheet['D36'] = b*10;  
    h = float (c1ent04.get());    sheet['C5'] = h
    pmin =  float (c1ent05.get())
    fiv = float (c1ent06.get())

    # Datos iniciales Parrilla sup
    rs = float (c1ent07.get())
    ds = h-rs;                       sheet['C6'] = ds; 
    Mnx = float (c2ent04.get());     sheet['C11'] = Mnx; 
    Mny = float (c2ent54.get());     sheet['D11'] = Mny; 
    Nbsx = float (c2ent08.get())
    dbpsx = f_dbp(Nbsx);             sheet['C16'] = dbpsx; 
    Nbsy = float (c2ent58.get())
    dbpsy = f_dbp(Nbsy);             sheet['D16'] = dbpsy; 

    # Datos iniciales Parrilla inf
    ri = float (c1ent08.get())
    di = h-ri;                      sheet['C7'] = di; 
    Mpx = float (c3ent04.get());    sheet['C23'] = Mpx; 
    Mpy = float (c3ent54.get());    sheet['D23'] = Mpy; 
    Nbix = float (c3ent08.get())
    dbpix = f_dbp(Nbix);            sheet['C28'] = dbpix; 
    Nbiy = float (c3ent58.get())
    dbpiy = f_dbp(Nbiy);            sheet['D28'] = dbpiy 

    # Datos iniciales diseño cortante simple
    r = max(rs, ri)
    d = h-r;                        sheet['C35'] = d*10; sheet['D35'] = d*10 
    Vux = float (c4ent04.get());    sheet['C37'] = Vux; 
    Vuy = float (c4ent54.get());    sheet['D37'] = Vuy; 

    # Acero mínimo
    Asmin = f_Asmin(pmin, b, h)    
    sheet['C14'] = "%.2f" %Asmin; sheet['D14'] = "%.2f" %Asmin; sheet['C26'] = "%.2f" %Asmin; sheet['D26'] = "%.2f" %Asmin 

    #Parrilla superior dir x - Diseño
    pcalsx = f_pcal(fc, fy, Mnx, b, ds);    sheet['C12'] = "%.4f" %pcalsx
    Ascalsx = f_Ascal(pcalsx, b, ds);       sheet['C13'] = "%.2f" %Ascalsx
    Asreqsx = f_Asreq(Ascalsx, Asmin);      sheet['C15'] = "%.2f" %Asreqsx
    Asbsx = f_Asb(Nbsx)
    Sepsx = f_sep(Asreqsx, Asbsx, b);       sheet['C17'] = "%.2f" %Sepsx
    dbpsx = f_dbp(Nbsx)
    Dissx = f_dist(Asreqsx, Asbsx, b, dbpsx)
    Ascolsx = f_Ascol(Asreqsx, Asbsx)
    Mnsx = f_Mn(fc, fy, Ascolsx, b, ds);    sheet['C19'] = "%.1f" %Mnsx
    ChMnsx = f_ChMn(Mnx,Mnsx);              sheet['C20'] = ChMnsx

    #Parrilla superior dir y - Diseño
    pcalsy = f_pcal(fc, fy, Mny, b, ds);    sheet['D12'] = "%.4f" %pcalsy
    Ascalsy = f_Ascal(pcalsy, b, ds);       sheet['D13'] = "%.2f" %Ascalsy
    Asreqsy = f_Asreq(Ascalsy, Asmin);      sheet['D15'] = "%.2f" %Asreqsy
    Asbsy = f_Asb(Nbsy)
    Sepsy = f_sep(Asreqsy, Asbsy, b);       sheet['D17'] = "%.2f" %Sepsy
    Dissy = f_dist(Asreqsy, Asbsy, b, Nbsy)
    Ascolsy = f_Ascol(Asreqsy, Asbsy)
    Mnsy = f_Mn(fc, fy, Ascolsy, b, ds);    sheet['D19'] = "%.1f" %Mnsy
    ChMnsy = f_ChMn(Mny,Mnsy);              sheet['D20'] = ChMnsy

    #Parrilla inferior dir x - Diseño
    pcalix = f_pcal(fc, fy, Mpx, b, di);    sheet['C24'] = "%.4f" %pcalix
    Ascalix = f_Ascal(pcalix, b, di);       sheet['C25'] = "%.2f" %Ascalix
    Asreqix = f_Asreq(Ascalix, Asmin);      sheet['C27'] = "%.2f" %Asreqix
    Asbix = f_Asb(Nbix)
    Sepix = f_sep(Asreqix, Asbix, b);       sheet['C29'] = "%.2f" %Sepix
    Disix = f_dist(Asreqix, Asbix, b, Nbix)
    Ascolix = f_Ascol(Asreqix, Asbix)
    Mnix = f_Mn(fc, fy, Ascolix, b, di);    sheet['C31'] = "%.1f" %Mnix
    ChMnix = f_ChMn(Mpx,Mnix);              sheet['C32'] = ChMnix

    #Parrilla inferior dir y - Diseño
    pcaliy = f_pcal(fc, fy, Mpy, b, di);    sheet['D24'] = "%.4f" %pcaliy
    Ascaliy = f_Ascal(pcaliy, b, di);       sheet['D25'] = "%.2f" %Ascaliy
    Asreqiy = f_Asreq(Ascaliy, Asmin);      sheet['D27'] = "%.2f" %Asreqiy
    Asbiy = f_Asb(Nbiy)
    Sepiy = f_sep(Asreqiy, Asbiy, b);       sheet['D29'] = "%.2f" %Sepiy
    Disiy = f_dist(Asreqiy, Asbiy, b, Nbiy)
    Ascoliy = f_Ascol(Asreqiy, Asbiy)
    Mniy = f_Mn(fc, fy, Ascoliy, b, di);    sheet['D31'] = "%.1f" %Mniy
    ChMniy = f_ChMn(Mpy,Mniy);              sheet['D32'] = ChMniy

    #Diseño a cortante
    Vcx = f_Vc(fiv, fc, b, d);              sheet['C38'] = "%.1f" %Vcx
    ChVcx = f_ChVc(Vux,Vcx);                sheet['C39'] = ChVcx
    Vcy = f_Vc(fiv, fc, b, d);              sheet['D38'] = "%.1f" %Vcy
    ChVcy = f_ChVc(Vuy,Vcy);                sheet['D39'] = ChVcy

    book.save ('D:\IEB_REPORTES\Slab Design.xlsx')

# Funcion borrar datos
def Borrar ():
    c1ent01.delete(0, 'end')
    c1ent02.delete(0, 'end')
    c1ent03.delete(0, 'end')
    c1ent04.delete(0, 'end')
    c1ent05.delete(0, 'end')
    c1ent06.delete(0, 'end')
    c1ent07.delete(0, 'end')
    c1ent08.delete(0, 'end')
    c1ent09.delete(0, 'end')

    c2ent04.delete(0, 'end');       c2ent54.delete(0, 'end')
    c2ent05.delete(0, 'end');       c2ent55.delete(0, 'end')
    c2ent06.delete(0, 'end');       c2ent56.delete(0, 'end')
    c2ent07.delete(0, 'end');       c2ent57.delete(0, 'end')
    c2ent08.delete(0, 'end');       c2ent58.delete(0, 'end')
    c2ent09.delete(0, 'end');       c2ent59.delete(0, 'end')
    c2ent13.delete(0, 'end');       c2ent63.delete(0, 'end')
    c2ent14.delete(0, 'end');       c2ent64.delete(0, 'end')
    c2ent15.delete(0, 'end');       c2ent65.delete(0, 'end')

    c3ent04.delete(0, 'end');       c3ent54.delete(0, 'end')
    c3ent05.delete(0, 'end');       c3ent55.delete(0, 'end')
    c3ent06.delete(0, 'end');       c3ent56.delete(0, 'end')
    c3ent07.delete(0, 'end');       c3ent57.delete(0, 'end')
    c3ent08.delete(0, 'end');       c3ent58.delete(0, 'end')
    c3ent09.delete(0, 'end');       c3ent59.delete(0, 'end')
    c3ent13.delete(0, 'end');       c3ent63.delete(0, 'end')
    c3ent14.delete(0, 'end');       c3ent64.delete(0, 'end')
    c3ent15.delete(0, 'end');       c3ent65.delete(0, 'end')

    c4ent04.delete(0, 'end');       c4ent54.delete(0, 'end')
    c4ent05.delete(0, 'end');       c4ent55.delete(0, 'end')
    c4ent06.delete(0, 'end');       c4ent56.delete(0, 'end')

# Recuadro 1 Datos iniciales
rec1 = LabelFrame(vent, text = '  Datos iniciales.  '); rec1.pack()
rec1.place(x=345, y=5, width=320, height=300)

c1tex01 = Label(rec1, text = "fc. Resistencia del concreto _ MPa"); c1tex01.pack()
c1tex01.place(x=10, y=10, width=210, height=20)
c1ent01 = Entry(rec1, justify=CENTER);                              c1ent01.place(x=230, y=10, width=80, height=20)

c1tex02 = Label(rec1, text = "fy. Fluencia del acero _ MPa");       c1tex02.pack()
c1tex02.place(x=10, y=40, width=210, height=20)
c1ent02 = Entry(rec1, justify=CENTER);                              c1ent02.place(x=230, y=40, width=80, height=20)

c1tex03 = Label(rec1, text = "b. Ancho de la sección _ cm");        c1tex03.pack()
c1tex03.place(x=10, y=70, width=210, height=20)
c1ent03 = Entry(rec1, justify=CENTER);                              c1ent03.place(x=230, y=70, width=80, height=20)

c1tex04 = Label(rec1, text = "h. Espesor de la losa _ cm");         c1tex04.pack()
c1tex04.place(x=10, y=100, width=210, height=20)
c1ent04 = Entry(rec1, justify=CENTER);                              c1ent04.place(x=230, y=100, width=80, height=20)

c1tex05 = Label(rec1, text = "pmín. Cuantía mínima.");              c1tex05.pack()
c1tex05.place(x=10, y=130, width=210, height=20)
c1ent05 = Entry(rec1, justify=CENTER);                              c1ent05.place(x=230, y=130, width=80, height=20)

c1tex06 = Label(rec1, text = "øv. Coeficiente de fricción.");        c1tex06.pack()
c1tex06.place(x=10, y=160, width=210, height=20)
c1ent06 = Entry(rec1, justify=CENTER);                              c1ent06.place(x=230, y=160, width=80, height=20)

c1tex07 = Label(rec1, text = "Recubrimiento parrilla sup _ cm");    c1tex07.pack()
c1tex07.place(x=10, y=190, width=210, height=20)
c1ent07 = Entry(rec1, justify=CENTER);                              c1ent07.place(x=230, y=190, width=80, height=20)

c1tex08 = Label(rec1, text = "Recubrimiento parrilla inf_ cm");     c1tex08.pack()
c1tex08.place(x=10, y=220, width=210, height=20)
c1ent08 = Entry(rec1, justify=CENTER);                              c1ent08.place(x=230, y=220, width=80, height=20)

c1tex09 = Label(rec1, text = "Asmin. Acero mínimo _ cm²");          c1tex09.pack()
c1tex09.place(x=10, y=250, width=210, height=20)
c1ent09 = Entry(rec1, justify=CENTER);                              c1ent09.place(x=230, y=250, width=80, height=20); c1ent09.config(bg="#ecf0f1")

# Recuadro 2 Parrilla superior

rec2 = LabelFrame(vent, text = '  Diseño parilla superior.  '); rec2.pack()
rec2.place(x=5, y=315, width=490, height=370)

c2tex01 = Label(rec2, text = "");                                             c2tex01.pack()
c2tex01.place(x=10, y=10, width=120, height=20)

c2tex02 = Label(rec2, text = "Dir longt _ Mxx (-)", font='Helvetica 7 bold'); c2tex02.pack()
c2tex02.place(x=230, y=10, width=120, height=20)

c2tex03 = Label(rec2, text = "Dir Trans _ Myy (-)", font='Helvetica 7 bold'); c2tex03.pack()
c2tex03.place(x=360, y=10, width=120, height=20)

c2tex04 = Label(rec2, text = "Momento de diseño _ kN.cm");          c2tex01.pack()
c2tex04.place(x=10, y=40, width=210, height=20)
c2ent04 = Entry(rec2, justify=CENTER);                              c2ent04.place(x=230, y=40, width=120, height=20)
c2ent54 = Entry(rec2, justify=CENTER);                              c2ent54.place(x=360, y=40, width=120, height=20)

c2tex05 = Label(rec2, text = "preq. Cuantía requerida.");           c2tex05.pack()
c2tex05.place(x=10, y=70, width=210, height=20)
c2ent05 = Entry(rec2, justify=CENTER);                              c2ent05.place(x=230, y=70, width=120, height=20); c2ent05.config(bg="#ecf0f1")
c2ent55 = Entry(rec2, justify=CENTER);                              c2ent55.place(x=360, y=70, width=120, height=20); c2ent55.config(bg="#ecf0f1")

c2tex06 = Label(rec2, text = "Asreq. Acero requerido _ cm²");       c2tex06.pack()
c2tex06.place(x=10, y=100, width=210, height=20)
c2ent06 = Entry(rec2, justify=CENTER);                              c2ent06.place(x=230, y=100, width=120, height=20); c2ent06.config(bg="#ecf0f1")
c2ent56 = Entry(rec2, justify=CENTER);                              c2ent56.place(x=360, y=100, width=120, height=20); c2ent56.config(bg="#ecf0f1")

c2tex07 = Label(rec2, text = "Ascol. Acero colocado _ cm²");        c2tex07.pack()
c2tex07.place(x=10, y=130, width=210, height=20)
c2ent07 = Entry(rec2, justify=CENTER);                              c2ent07.place(x=230, y=130, width=120, height=20); c2ent07.config(bg="#ecf0f1")
c2ent57 = Entry(rec2, justify=CENTER);                              c2ent57.place(x=360, y=130, width=120, height=20); c2ent57.config(bg="#ecf0f1")

c2tex08 = Label(rec2, text = "N° de barra a usar.");                c2tex08.pack()
c2tex08.place(x=10, y=160, width=210, height=20)
c2ent08 = Entry(rec2, justify=CENTER);                              c2ent08.place(x=230, y=160, width=120, height=20)
c2ent58 = Entry(rec2, justify=CENTER);                              c2ent58.place(x=360, y=160, width=120, height=20)

c2tex09 = Label(rec2, text = "Distribución del refuerzo.");         c2tex09.pack()
c2tex09.place(x=10, y=190, width=210, height=20)
c2ent09 = Entry(rec2, justify=CENTER);                              c2ent09.place(x=230, y=190, width=120, height=20); c2ent09.config(bg="#ecf0f1") #, font='Arial 9')
c2ent59 = Entry(rec2, justify=CENTER);                              c2ent59.place(x=360, y=190, width=120, height=20); c2ent59.config(bg="#ecf0f1") #, font='Arial 9')

c2tex10 = Label(rec2, text = "Chequeo momento nominal.");           c2tex10.pack()
c2tex10.place(x=10, y=230, width=210, height=20)

c2tex11 = Label(rec2, text = "Mn _ Dir x", font='Helvetica 7 bold'); c2tex11.pack()
c2tex11.place(x=230, y=230, width=120, height=20)

c2tex12 = Label(rec2, text = "Mn _ Dir y", font='Helvetica 7 bold'); c2tex12.pack()
c2tex12.place(x=360, y=230, width=120, height=20)

c2tex13 = Label(rec2, text = "Ascol. Acero colocado real _ cm²");   c2tex13.pack()
c2tex13.place(x=10, y=260, width=210, height=20)
c2ent13 = Entry(rec2, justify=CENTER);                              c2ent13.place(x=230, y=260, width=120, height=20); c2ent13.config(bg="#ecf0f1")
c2ent63 = Entry(rec2, justify=CENTER);                              c2ent63.place(x=360, y=260, width=120, height=20); c2ent63.config(bg="#ecf0f1")

c2tex14 = Label(rec2, text = "øMn. Momento nominal _ kN.cm");       c2tex14.pack()
c2tex14.place(x=10, y=290, width=210, height=20)
c2ent14 = Entry(rec2, justify=CENTER);                              c2ent14.place(x=230, y=290, width=120, height=20); c2ent14.config(bg="#ecf0f1")
c2ent64 = Entry(rec2, justify=CENTER);                              c2ent64.place(x=360, y=290, width=120, height=20); c2ent64.config(bg="#ecf0f1")

c2tex15 = Label(rec2, text = "Mu. < øMn.");                         c2tex15.pack()
c2tex15.place(x=10, y=320, width=210, height=20)
c2ent15 = Entry(rec2, justify=CENTER);                              c2ent15.place(x=230, y=320, width=120, height=20); c2ent15.config(bg="#ecf0f1")
c2ent65 = Entry(rec2, justify=CENTER);                              c2ent65.place(x=360, y=320, width=120, height=20); c2ent65.config(bg="#ecf0f1")

# Recuadro 3. Parrilla inferior

rec3 = LabelFrame(vent, text = '  Diseño parrilla inferior.  '); rec3.pack()
rec3.place(x=500, y=315, width=490, height=370)

c3tex01 = Label(rec3, text = "");                                             c3tex01.pack()
c3tex01.place(x=10, y=10, width=120, height=20)

c3tex02 = Label(rec3, text = "Dir longt _ Mxx (+)", font='Helvetica 7 bold'); c3tex02.pack()
c3tex02.place(x=230, y=10, width=120, height=20)

c3tex03 = Label(rec3, text = "Dir Trans _ Myy (+)", font='Helvetica 7 bold'); c3tex03.pack()
c3tex03.place(x=360, y=10, width=120, height=20)

c3tex04 = Label(rec3, text = "Momento de diseño _ kN.cm");          c3tex01.pack()
c3tex04.place(x=10, y=40, width=210, height=20)
c3ent04 = Entry(rec3, justify=CENTER);                              c3ent04.place(x=230, y=40, width=120, height=20)
c3ent54 = Entry(rec3, justify=CENTER);                              c3ent54.place(x=360, y=40, width=120, height=20)

c3tex05 = Label(rec3, text = "preq. Cuantía requerida.");           c3tex05.pack()
c3tex05.place(x=10, y=70, width=210, height=20)
c3ent05 = Entry(rec3, justify=CENTER);                              c3ent05.place(x=230, y=70, width=120, height=20); c3ent05.config(bg="#ecf0f1")
c3ent55 = Entry(rec3, justify=CENTER);                              c3ent55.place(x=360, y=70, width=120, height=20); c3ent55.config(bg="#ecf0f1")

c3tex06 = Label(rec3, text = "Asreq. Acero requerido _ cm²");       c3tex06.pack()
c3tex06.place(x=10, y=100, width=210, height=20)
c3ent06 = Entry(rec3, justify=CENTER);                              c3ent06.place(x=230, y=100, width=120, height=20); c3ent06.config(bg="#ecf0f1")
c3ent56 = Entry(rec3, justify=CENTER);                              c3ent56.place(x=360, y=100, width=120, height=20); c3ent56.config(bg="#ecf0f1")

c3tex07 = Label(rec3, text = "Ascol. Acero colocado _ cm²");        c3tex07.pack()
c3tex07.place(x=10, y=130, width=210, height=20)
c3ent07 = Entry(rec3, justify=CENTER);                              c3ent07.place(x=230, y=130, width=120, height=20); c3ent07.config(bg="#ecf0f1")
c3ent57 = Entry(rec3, justify=CENTER);                              c3ent57.place(x=360, y=130, width=120, height=20); c3ent57.config(bg="#ecf0f1")

c3tex08 = Label(rec3, text = "N° de barra a usar.");                c3tex08.pack()
c3tex08.place(x=10, y=160, width=210, height=20)
c3ent08 = Entry(rec3, justify=CENTER);                              c3ent08.place(x=230, y=160, width=120, height=20)
c3ent58 = Entry(rec3, justify=CENTER);                              c3ent58.place(x=360, y=160, width=120, height=20)

c3tex09 = Label(rec3, text = "Distribución del refuerzo.");         c3tex09.pack()
c3tex09.place(x=10, y=190, width=210, height=20)
c3ent09 = Entry(rec3, justify=CENTER);                              c3ent09.place(x=230, y=190, width=120, height=20); c3ent09.config(bg="#ecf0f1") #, font='Arial 9')
c3ent59 = Entry(rec3, justify=CENTER);                              c3ent59.place(x=360, y=190, width=120, height=20); c3ent59.config(bg="#ecf0f1") #, font='Arial 9')

c3tex10 = Label(rec3, text = "Chequeo momento nominal.");            c3tex10.pack()
c3tex10.place(x=10, y=230, width=210, height=20)

c3tex11 = Label(rec3, text = "Mn _ Dir x", font='Helvetica 7 bold'); c3tex11.pack()
c3tex11.place(x=230, y=230, width=120, height=20)

c3tex12 = Label(rec3, text = "Mn _ Dir y", font='Helvetica 7 bold'); c3tex12.pack()
c3tex12.place(x=360, y=230, width=120, height=20)

c3tex13 = Label(rec3, text = "Ascol. Acero colocado real _ cm²");   c3tex13.pack()
c3tex13.place(x=10, y=260, width=210, height=20)
c3ent13 = Entry(rec3, justify=CENTER);                              c3ent13.place(x=230, y=260, width=120, height=20); c3ent13.config(bg="#ecf0f1")
c3ent63 = Entry(rec3, justify=CENTER);                              c3ent63.place(x=360, y=260, width=120, height=20); c3ent63.config(bg="#ecf0f1")

c3tex14 = Label(rec3, text = "øMn. Momento nominal _ kN.cm");       c3tex14.pack()
c3tex14.place(x=10, y=290, width=210, height=20)
c3ent14 = Entry(rec3, justify=CENTER);                              c3ent14.place(x=230, y=290, width=120, height=20); c3ent14.config(bg="#ecf0f1")
c3ent64 = Entry(rec3, justify=CENTER);                              c3ent64.place(x=360, y=290, width=120, height=20); c3ent64.config(bg="#ecf0f1")

c3tex15 = Label(rec3, text = "Mu. < øMn.");                         c3tex15.pack()
c3tex15.place(x=10, y=320, width=210, height=20)
c3ent15 = Entry(rec3, justify=CENTER);                              c3ent15.place(x=230, y=320, width=120, height=20); c3ent15.config(bg="#ecf0f1")
c3ent65 = Entry(rec3, justify=CENTER);                              c3ent65.place(x=360, y=320, width=120, height=20); c3ent65.config(bg="#ecf0f1")

# Recuadro 4 Cortante

rec4 = LabelFrame(vent, text = '  Cortante simple.  '); rec4.pack()
rec4.place(x=5, y=690, width=490, height=150)

c4tex01 = Label(rec4, text = "Chequeo a cortante simple.");           c4tex01.pack()
c4tex01.place(x=10, y=10, width=210, height=20)

c4tex02 = Label(rec4, text = "Vc _ Dir x", font='Helvetica 7 bold');  c4tex02.pack()
c4tex02.place(x=230, y=10, width=120, height=20)

c4tex03 = Label(rec4, text = "Vc _ Dir Vy", font='Helvetica 7 bold'); c4tex03.pack()
c4tex03.place(x=360, y=10, width=120, height=20)

c4tex04 = Label(rec4, text = "Vu. Cortante de diseño _ kN");        c4tex04.pack()
c4tex04.place(x=10, y=40, width=210, height=20)
c4ent04 = Entry(rec4, justify=CENTER);                              c4ent04.place(x=230, y=40, width=120, height=20)
c4ent54 = Entry(rec4, justify=CENTER);                              c4ent54.place(x=360, y=40, width=120, height=20)

c4tex05 = Label(rec4, text = "øVc. Resitencia del concreto _ kN");  c4tex05.pack()
c4tex05.place(x=10, y=70, width=210, height=20)
c4ent05 = Entry(rec4, justify=CENTER);                              c4ent05.place(x=230, y=70, width=120, height=20); c4ent05.config(bg="#ecf0f1")
c4ent55 = Entry(rec4, justify=CENTER);                              c4ent55.place(x=360, y=70, width=120, height=20); c4ent55.config(bg="#ecf0f1")

c4tex06 = Label(rec4, text = "Vu. < øVc.");                         c4tex06.pack()
c4tex06.place(x=10, y=100, width=210, height=20)
c4ent06 = Entry(rec4, justify=CENTER);                              c4ent06.place(x=230, y=100, width=120, height=20); c4ent06.config(bg="#ecf0f1")
c4ent56 = Entry(rec4, justify=CENTER);                              c4ent56.place(x=360, y=100, width=120, height=20); c4ent56.config(bg="#ecf0f1")

# Recuadro 5 botones

rec5 = LabelFrame(vent, text = '  Opciones.  '); rec5.pack()
rec5.place(x=500, y=725, width=490, height=60)

bot1 = Button(rec5, text = 'Calcular', font='Helvetica 8 bold', command=Calcular);      bot1.pack()
bot1.place(x=32.5, y=10, width=120, height=20)

bot2 = Button(rec5, text = 'Guardar .xls', font='Helvetica 8 bold', command=Guardar);   bot2.pack()
bot2.place(x=185, y=10, width=120, height=20)

bot3 = Button(rec5, text = 'Borrar', font='Helvetica 8 bold', command=Borrar );         bot3.pack()
bot3.place(x=337.5, y=10, width=120, height=20)

label = Label(vent, text = "wilson.taimalc@gmail.com - 2023", font='Arial 7'); label.pack()
label.place(x=500, y=795, width=490, height=10)

vent.mainloop()

